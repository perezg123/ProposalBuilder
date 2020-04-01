import os

from flask import Flask, request, jsonify, url_for, send_from_directory, render_template, session
from openpyxl import load_workbook
from tablib import Dataset
from werkzeug.utils import secure_filename

app = Flask(__name__)
ALLOWED_EXTENSIONS = {'xlsx'}
REMOVE_TABS = ['Cover Sheet', 'Index', 'General Info', 'DataSet', 'Changes']
categories = []
products = []
app.secret_key = "soin342in*^Qub"
global filename

@app.route('/favicon.ico')
def favicon():
    return send_from_directory(os.path.join(app.root_path, 'static'), 'favicon.ico', mimetype='image/vnd.microsoft.icon')

@app.route('/', methods=['GET', 'POST'])
def file_select():
    return '''
    <!doctype html>
    <title>Load Product File</title>
    <h1>Load Product File</h1>
    <form method="post" action="load" enctype="multipart/form-data">
      <p><input type="file" name="file">
         <input type="submit" value="Upload">
    </form>
    '''

@app.route('/load', methods=['POST'])
def loadfile():
    global filename
    # I used form data type which means there is a
    # "Content-Type: application/x-www-form-urlencoded"
    # header in my request
#    raw_data = request.files['file'].read()  # In form data, I used "file" as key.
#    dataset = Dataset().load(raw_data)
#    return jsonify(dataset.export('json'))
    workbook = load_workbook(request.files['file'], read_only=True)
    filename = request.files['file']
    products = workbook.sheetnames
    for sheet in products:
        if sheet in REMOVE_TABS:
            pass
        else:
            categories.append(sheet)
    return render_template('index.html', filename=request.files['file'], categories = categories)

@app.route('/product', methods=['POST'])
def show_product():
    global filename
    unit_header = False
    workbook = load_workbook('/Users/perezg/Downloads/' + str(filename.filename), read_only=True)
    dict = request.form
    product = dict['product']
    product_data = {}
    for sheet in workbook:
        print(sheet.title)
    print('Reading Rows...')
    sheet = workbook[product]
    for row in range(30, sheet.max_row):
        if (sheet['A' + str(row)].value == 'UNIT') and (sheet['B' + str(row)].value == 'SKU'):
            print (sheet['A' + str(row)].value + " : " + sheet['B' + str(row)].value)
            unit_header = True
            unit = sheet['A' + str(row + 1)].value
            product_data.setdefault(unit, {})
            continue
        if unit_header:
            sku = sheet['B' + str(row)].value
            descr = sheet['C' + str(row)].value
            price = sheet['D' + str(row)].value
            oneyr = sheet['E' + str(row)].value
            twoyr = sheet['F' + str(row)].value
            threeyr = sheet['G' + str(row)].value
            fouryr = sheet['H' + str(row)].value
            fiveyr = sheet['I' + str(row)].value
            comments = sheet['J' + str(row)].value
            cat = sheet['K' + str(row)].value
            product_data[unit] = ({'sku': '', 'descr': '', 'price': '', 'oneyr': '', 'twoyr': '', 'threeyr': '',
                                   'fouryr': '', 'fiveyr': '', 'comments': '', 'cat': ''})
 #       print (sheet['A' + str(row)].value, sheet['B' + str(row)].value)
    return render_template('index.html', filename=request.files['file'], categories = categories, productdata = product_data)

if __name__ == '__main__':
    app.run(debug=True)