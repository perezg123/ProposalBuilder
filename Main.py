import os
import json
from openpyxl import Workbook, load_workbook
import datetime
from dataclasses import dataclass
from flask import Flask, request, render_template, url_for, redirect, jsonify, flash, sessions
from werkzeug.utils import secure_filename

from classes import Product

global unit_header
global workbook
global sheet

app = Flask(__name__)
app.secret_key = "soin342in*^Qub"
ALLOWED_EXTENSIONS = {'xlsx'}
UPLOAD_FOLDER = "/Users/perezg/Public/"
REMOVE_TABS = ['Cover Sheet', 'Index', 'General Info', 'DataSet', 'Changes']
categories = []
products = []

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        # check if the post request has the file part
        if 'file' not in request.files:
            flash('No file part')
            return redirect(request.url)
        file = request.files['file']
        # if user does not select file, browser also
        # submit a empty part without filename
        if file.filename == '':
            flash('No selected file')
            return redirect(request.url)
        if file and allowed_file(file.filename):
            file = request.files['file']
            file.save(os.path.join(UPLOAD_FOLDER, file.filename))
            workbook = load_workbook(file.filename, read_only=True)
            products = workbook.sheetnames
            for sheet in products:
                if sheet in REMOVE_TABS:
                    pass
                else:
                    categories.append(sheet)
            return render_template('index.html', filename=filename, categories = categories)
    return '''
    <!doctype html>
    <title>Upload new File</title>
    <h1>Upload new File</h1>
    <form method=post enctype=multipart/form-data>
      <p><input type=file name=file>
         <input type=submit value=Upload>
    </form>
    '''


@app.route("/main")
def mainPage():
    return render_template('index.html')

@app.route("/uploadfile", methods=['GET', 'POST'])
def uploadfile():
    if request.method == 'POST':
        data=request.form.to_dict()
        filename = request.files['filename']
        print (data)
        if 'category' in data:
            workbook = load_workbook(filename=secure_filename(request.args.get('file')), read_only=True)
            sheet = workbook.get_sheet_by_name(request.args.get('category'))
            print(request.args.get('category'))
            product_data = {}
            print('Reading Rows...')
            for row in range(2,sheet.max_row + 1):
                if (sheet['A' + str(row)].value == 'UNIT') and (sheet['A' + str(row)].value == 'SKU'):
                    unit_header = True
                    unit = sheet['A' + str(row + 1)].value
                    product_data.setdefault(unit,{})
                    continue
                if unit_header:
                    sku = sheet['B' + str(row)].value
                    descr = sheet['C' + str(row)].value
                    price = sheet['D' + str(row)].value
                    oneyr = sheet['E' + str(row)].value
                    twoyr = sheet['F' + str(row)].value
                    threeyr = sheet['G' + str(row)].value
                    fouryr = sheet['H' + str(row)].value
                    fiveyr = sheet['I' + str(row)].value
                    comments = sheet['J' + str(row)].value
                    cat = sheet['K' + str(row)].value
                    product_data[unit] = ({'sku':'','descr':'','price':'','oneyr':'','twoyr':'','threeyr':'',
                                                   'fouryr':'','fiveyr':'','comments':'','cat':''})
        if filename and allowed_file(filename.filename):
            filename = secure_filename(filename.filename)
#            filename=data['filename']
            workbook = load_workbook(filename=filename, read_only=True)
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            products = workbook.sheetnames
            for sheet in products:
                if sheet in REMOVE_TABS:
                    pass
                else:
                    categories.append(sheet)
            return render_template('index.html', filename=request.args.get('file'), categories = categories)

#    return render_template('upload.html')
    if request.method == 'GET':
        return '''
            <!doctype html>
            <title>Upload an excel file</title>
            <h1>Excel file upload (csv, tsv, csvz, tsvz only)</h1>
            <form action="uploadfile" method="post" enctype="multipart/form-data">
            <p><input type=file name=file><input type=submit value='Load Data'>
            </form>
            '''
    else:
        return ('Invalid Method Received')

@app.route("/handleFileUpload", methods=['POST'])
def handleFileUpload():
    if 'file' in request.files:
        input_file = request.files['file']
        if input_file.filename != '':
            input_file.save(os.path.join('/tmp', input_file.filename))
        workbook = load_workbook(filename=input_file.filename, read_only=True)
        sheet = workbook.active
        categories = workbook.sheetnames
        print(categories)
        return render_template('fileform.html', categories)
    return redirect(url_for('fileFrontPage'))


if __name__ == '__main__':
    app.run()
